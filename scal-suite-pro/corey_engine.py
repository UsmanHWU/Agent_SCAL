"""
corey_engine.py
---------------
SCAL Simulation Engine: Relative Permeability & Capillary Pressure
Supports: Corey, Modified Corey (LET), Burdine, Chierici, and Brooks-Corey methods.
Outputs: Pandas DataFrame, Eclipse .inc files, Excel workbook with embedded chart.

References:
  - Dake, L.P., Fundamentals of Reservoir Engineering
  - Corey, A.T. (1954) - The interrelation between gas and oil relative permeabilities
  - LET model: Lomeland, Ebeltoft, Thomas (2005) SPE-89352
  - Burdine, N.T. (1953) Trans. AIME
  - Brooks & Corey (1964) Hydraulic Properties of Porous Media
  - Chierici, G.L. (1984) Novel relations for drainage and imbibition curves
"""

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
import textwrap
from datetime import datetime

# ─────────────────────────────────────────────────────────────────────────────
# SATURATION NORMALISATION
# ─────────────────────────────────────────────────────────────────────────────

def normalise(sw, swc, s_res):
    """Return normalised water saturation Swn ∈ [0, 1]."""
    denom = 1.0 - swc - s_res
    if denom <= 0:
        raise ValueError(f"Swc ({swc}) + S_res ({s_res}) ≥ 1.0 — unphysical saturation table.")
    return np.clip((sw - swc) / denom, 0.0, 1.0)


def sw_array(swc, s_res, n_points=30):
    """Return array of n_points Sw values spanning [Swc, 1 - S_res]."""
    return np.linspace(swc, 1.0 - s_res, n_points)


# ─────────────────────────────────────────────────────────────────────────────
# METHOD 1 — COREY (1954)
# ─────────────────────────────────────────────────────────────────────────────

def corey(system, swc, s_res, krw_end, kr_phase_end, nw, n_phase, include_pc,
          pc_entry=1.0, lambda_pc=2.0, uncert_pct=0.0, n_points=30):
    """
    Classical Corey power-law relative permeability.
      krw  = krw_end  * Swn ^ nw
      kro  = krw_phase_end * (1 - Swn) ^ n_phase    [Oil-Water]
      krg  = kr_phase_end * (1 - Swn) ^ n_phase     [Gas-Water]
    Capillary pressure: Brooks-Corey Pc = pc_entry * Swn^(-1/lambda_pc)
    """
    if uncert_pct > 0:
        rng = np.random.default_rng()
        swc = float(np.clip(swc * rng.uniform(1 - uncert_pct, 1 + uncert_pct), 0.05, 0.45))
        s_res = float(np.clip(s_res * rng.uniform(1 - uncert_pct, 1 + uncert_pct), 0.05, 0.45))

    sw = sw_array(swc, s_res, n_points)
    swn = normalise(sw, swc, s_res)

    krw = krw_end * swn ** nw
    kr_p = kr_phase_end * (1.0 - swn) ** n_phase

    if include_pc:
        # Avoid division by zero at swn=0
        swn_safe = np.where(swn < 1e-6, 1e-6, swn)
        pc = pc_entry * swn_safe ** (-1.0 / lambda_pc)
        pc = np.clip(pc, 0.0, 100.0)
    else:
        pc = np.zeros_like(sw)

    phase_label = "Krg" if system == "Gas-Water" else "Kro"
    return pd.DataFrame({"Sw": sw, "Krw": krw, phase_label: kr_p, "Pc": pc})


# ─────────────────────────────────────────────────────────────────────────────
# METHOD 2 — LET / MODIFIED COREY (Lomeland-Ebeltoft-Thomas, SPE-89352, 2005)
# ─────────────────────────────────────────────────────────────────────────────

def let_kr(system, swc, s_res, krw_end, kr_phase_end,
           Lw, Ew, Tw, Lp, Ep, Tp,
           include_pc, pc_entry=1.0, lambda_pc=2.0,
           uncert_pct=0.0, n_points=30):
    """
    LET (Modified Corey) model:
      krw  = krw_end  * Swn^L  /  [Swn^L  + E*(1-Swn)^T]
      kro  = krw_phase_end * (1-Swn)^Lp / [(1-Swn)^Lp + Ep*Swn^Tp]
    L, E, T are shape parameters for water and phase respectively.
    Provides greater flexibility for S-shaped curves seen in heterogeneous rocks.
    """
    if uncert_pct > 0:
        rng = np.random.default_rng()
        swc = float(np.clip(swc * rng.uniform(1 - uncert_pct, 1 + uncert_pct), 0.05, 0.45))
        s_res = float(np.clip(s_res * rng.uniform(1 - uncert_pct, 1 + uncert_pct), 0.05, 0.45))

    sw = sw_array(swc, s_res, n_points)
    swn = normalise(sw, swc, s_res)

    # Water LET
    denom_w = swn ** Lw + Ew * (1.0 - swn) ** Tw
    krw = krw_end * np.where(denom_w < 1e-15, 0.0, swn ** Lw / denom_w)

    # Phase (oil or gas) LET
    swn_inv = 1.0 - swn
    denom_p = swn_inv ** Lp + Ep * swn ** Tp
    kr_p = kr_phase_end * np.where(denom_p < 1e-15, 0.0, swn_inv ** Lp / denom_p)

    if include_pc:
        swn_safe = np.where(swn < 1e-6, 1e-6, swn)
        pc = pc_entry * swn_safe ** (-1.0 / lambda_pc)
        pc = np.clip(pc, 0.0, 100.0)
    else:
        pc = np.zeros_like(sw)

    phase_label = "Krg" if system == "Gas-Water" else "Kro"
    return pd.DataFrame({"Sw": sw, "Krw": krw, phase_label: kr_p, "Pc": pc})


# ─────────────────────────────────────────────────────────────────────────────
# METHOD 3 — BURDINE (1953)
# ─────────────────────────────────────────────────────────────────────────────

def burdine(system, swc, s_res, krw_end, kr_phase_end, lambda_b,
            include_pc, pc_entry=1.0, uncert_pct=0.0, n_points=30):
    """
    Burdine model derived from capillary tube bundle theory:
      krw  = krw_end  * Swn^2 * [1 - (1 - Swn^(1/lambda))^lambda]   (approx)
      kro  = kr_phase_end * (1-Swn)^2 * [1 - Swn^(1/lambda)]^lambda
    Here we use the standard pore-size-distribution form with exponent lambda.
    """
    if uncert_pct > 0:
        rng = np.random.default_rng()
        swc = float(np.clip(swc * rng.uniform(1 - uncert_pct, 1 + uncert_pct), 0.05, 0.45))
        s_res = float(np.clip(s_res * rng.uniform(1 - uncert_pct, 1 + uncert_pct), 0.05, 0.45))

    sw = sw_array(swc, s_res, n_points)
    swn = normalise(sw, swc, s_res)

    # Burdine integral approximation using Brooks-Corey pore distribution
    exp = (2.0 + lambda_b) / lambda_b
    krw = krw_end * swn ** exp
    kr_p = kr_phase_end * (1.0 - swn) ** 2 * (1.0 - swn ** ((2.0 + lambda_b) / lambda_b))

    if include_pc:
        swn_safe = np.where(swn < 1e-6, 1e-6, swn)
        pc = pc_entry * swn_safe ** (-1.0 / lambda_b)
        pc = np.clip(pc, 0.0, 100.0)
    else:
        pc = np.zeros_like(sw)

    phase_label = "Krg" if system == "Gas-Water" else "Kro"
    return pd.DataFrame({"Sw": sw, "Krw": krw, phase_label: kr_p, "Pc": pc})


# ─────────────────────────────────────────────────────────────────────────────
# METHOD 4 — BROOKS-COREY (1964)
# ─────────────────────────────────────────────────────────────────────────────

def brooks_corey(system, swc, s_res, krw_end, kr_phase_end, lambda_bc,
                 include_pc, pc_entry=1.0, uncert_pct=0.0, n_points=30):
    """
    Brooks-Corey (1964) model with pore-size distribution index lambda_bc:
      krw  = krw_end  * Swn^((2 + 3*lambda) / lambda)
      kro  = kr_phase_end * (1-Swn)^2 * (1 - Swn^((2+lambda)/lambda))
      Pc   = pc_entry * Swn^(-1/lambda)
    lambda_bc typical range: 0.2 (broad) to 7.0 (narrow/well-sorted)
    """
    if uncert_pct > 0:
        rng = np.random.default_rng()
        swc = float(np.clip(swc * rng.uniform(1 - uncert_pct, 1 + uncert_pct), 0.05, 0.45))
        s_res = float(np.clip(s_res * rng.uniform(1 - uncert_pct, 1 + uncert_pct), 0.05, 0.45))

    sw = sw_array(swc, s_res, n_points)
    swn = normalise(sw, swc, s_res)

    nw_exp = (2.0 + 3.0 * lambda_bc) / lambda_bc
    krw = krw_end * swn ** nw_exp
    kr_p = kr_phase_end * (1.0 - swn) ** 2 * (1.0 - swn ** ((2.0 + lambda_bc) / lambda_bc))

    if include_pc:
        swn_safe = np.where(swn < 1e-6, 1e-6, swn)
        pc = pc_entry * swn_safe ** (-1.0 / lambda_bc)
        pc = np.clip(pc, 0.0, 100.0)
    else:
        pc = np.zeros_like(sw)

    phase_label = "Krg" if system == "Gas-Water" else "Kro"
    return pd.DataFrame({"Sw": sw, "Krw": krw, phase_label: kr_p, "Pc": pc})


# ─────────────────────────────────────────────────────────────────────────────
# METHOD 5 — CHIERICI (1984)
# ─────────────────────────────────────────────────────────────────────────────

def chierici(system, swc, s_res, krw_end, kr_phase_end,
             aw, bw, ap, bp,
             include_pc, pc_entry=1.0, lambda_pc=2.0,
             uncert_pct=0.0, n_points=30):
    """
    Chierici (1984) exponential model — well-suited for fractured/vuggy carbonates:
      krw  = krw_end  * exp(-aw / Swn^bw)
      kro  = kr_phase_end * exp(-ap / (1-Swn)^bp)
    Avoids crossover by construction. aw, bw, ap, bp are curve-fit parameters
    typically obtained from laboratory SCAL measurements.
    """
    if uncert_pct > 0:
        rng = np.random.default_rng()
        swc = float(np.clip(swc * rng.uniform(1 - uncert_pct, 1 + uncert_pct), 0.05, 0.45))
        s_res = float(np.clip(s_res * rng.uniform(1 - uncert_pct, 1 + uncert_pct), 0.05, 0.45))

    sw = sw_array(swc, s_res, n_points)
    swn = normalise(sw, swc, s_res)

    swn_safe = np.where(swn < 1e-6, 1e-6, swn)
    swn_inv_safe = np.where((1.0 - swn) < 1e-6, 1e-6, 1.0 - swn)

    krw = krw_end * np.exp(-aw / swn_safe ** bw)
    kr_p = kr_phase_end * np.exp(-ap / swn_inv_safe ** bp)

    if include_pc:
        pc = pc_entry * swn_safe ** (-1.0 / lambda_pc)
        pc = np.clip(pc, 0.0, 100.0)
    else:
        pc = np.zeros_like(sw)

    phase_label = "Krg" if system == "Gas-Water" else "Kro"
    return pd.DataFrame({"Sw": sw, "Krw": krw, phase_label: kr_p, "Pc": pc})


# ─────────────────────────────────────────────────────────────────────────────
# ECLIPSE .INC FILE GENERATOR
# ─────────────────────────────────────────────────────────────────────────────

def export_to_eclipse_inc(df, system, method, swc, s_res, include_pc,
                           filename="SCAL_FUNCTIONS.INC"):
    """
    Writes Eclipse-format .INC file.
    For Oil-Water: SWOF keyword (Sw Krw Kro Pc)
    For Gas-Water: SGOF keyword (Sg Krg Krw Pc), converted from Sw basis
    Conforms to E100/E300 simulator syntax requirements.
    """
    phase_col = "Krg" if system == "Gas-Water" else "Kro"
    now = datetime.now().strftime("%Y-%m-%d %H:%M")

    header = textwrap.dedent(f"""\
    -- ============================================================
    -- Eclipse SCAL Include File
    -- Generated by SCAL Digital Assistant Suite
    -- Method  : {method}
    -- System  : {system}
    -- Date    : {now}
    -- Swc     : {swc:.4f}   S_res : {s_res:.4f}
    -- Pc incl : {"YES" if include_pc else "NO"}
    -- ============================================================
    """)

    if system == "Oil-Water":
        keyword = "SWOF"
        lines = [header, f"{keyword}\n--  Sw          Krw         Kro         Pc (bar)\n"]
        for _, row in df.iterrows():
            sw  = row["Sw"]
            krw = row["Krw"]
            kro = row[phase_col]
            pc  = row["Pc"]
            lines.append(f"   {sw:10.6f}  {krw:10.6f}  {kro:10.6f}  {pc:10.6f}")
        lines.append("/\n")

    else:  # Gas-Water — convert Sw to Sg; flip ordering (Eclipse SGOF expects Sg ascending)
        keyword = "SGOF"
        sg_col = 1.0 - df["Sw"]
        df_out = df.copy()
        df_out["Sg"] = sg_col
        df_out = df_out.sort_values("Sg").reset_index(drop=True)
        # In SGOF: Sg Krg Krw Pc
        lines = [header, f"{keyword}\n--  Sg          Krg         Krw         Pc (bar)\n"]
        for _, row in df_out.iterrows():
            sg  = row["Sg"]
            krg = row[phase_col]
            krw = row["Krw"]
            pc  = row["Pc"]
            lines.append(f"   {sg:10.6f}  {krg:10.6f}  {krw:10.6f}  {pc:10.6f}")
        lines.append("/\n")

    with open(filename, "w") as f:
        f.write("\n".join(lines))

    return filename


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL WORKBOOK WITH EMBEDDED CHART
# ─────────────────────────────────────────────────────────────────────────────

def export_to_excel(df, system, method, filename="SCAL_Report.xlsx"):
    """
    Saves SCAL data to an Excel workbook with a native line chart
    and a metadata summary sheet.
    """
    wb = Workbook()

    # ── Sheet 1: Raw data ──
    ws_data = wb.active
    ws_data.title = "SCAL_Data"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws_data.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        for c_idx, val in enumerate(row, start=1):
            ws_data.cell(row=r_idx, column=c_idx, value=round(float(val), 6))

    n_rows = len(df)

    # Chart — Kr curves
    chart = LineChart()
    chart.title = f"Relative Permeability — {method} ({system})"
    chart.x_axis.title = "Water Saturation Sw"
    chart.y_axis.title = "Relative Permeability (fraction)"
    chart.y_axis.scaling.min = 0.0
    chart.y_axis.scaling.max = 1.0
    chart.width = 20
    chart.height = 14

    kr_data = Reference(ws_data, min_col=2, max_col=3, min_row=1, max_row=n_rows + 1)
    cats    = Reference(ws_data, min_col=1, min_row=2, max_row=n_rows + 1)
    chart.add_data(kr_data, titles_from_data=True)
    chart.set_categories(cats)
    ws_data.add_chart(chart, "F2")

    # Chart — Pc
    if "Pc" in df.columns and df["Pc"].sum() > 0:
        chart_pc = LineChart()
        chart_pc.title = "Capillary Pressure Pc vs Sw"
        chart_pc.x_axis.title = "Water Saturation Sw"
        chart_pc.y_axis.title = "Pc (bar)"
        chart_pc.width = 20
        chart_pc.height = 14

        pc_data = Reference(ws_data, min_col=4, max_col=4, min_row=1, max_row=n_rows + 1)
        chart_pc.add_data(pc_data, titles_from_data=True)
        chart_pc.set_categories(cats)
        ws_data.add_chart(chart_pc, "F28")

    # ── Sheet 2: Metadata ──
    ws_meta = wb.create_sheet("Metadata")
    ws_meta.column_dimensions["A"].width = 35
    ws_meta.column_dimensions["B"].width = 35
    meta_rows = [
        ("Property", "Value"),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Method", method),
        ("Fluid System", system),
        ("Swc", f"{df['Sw'].min():.4f}"),
        ("1 - S_res", f"{df['Sw'].max():.4f}"),
        ("Max Krw", f"{df['Krw'].max():.4f}"),
        ("Max Kr_phase", f"{df.iloc[:, 2].max():.4f}"),
        ("Data Points", str(n_rows)),
    ]
    for r_idx, (k, v) in enumerate(meta_rows, start=1):
        ws_meta.cell(row=r_idx, column=1, value=k).font = Font(bold=(r_idx == 1))
        ws_meta.cell(row=r_idx, column=2, value=v)

    wb.save(filename)
    return filename
